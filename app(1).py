import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import re
import os
import csv
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# ── Palette ──────────────────────────────────────────────────────────────────
BG       = "#FFFFFF"
BG2      = "#F5F5F5"
FG       = "#0A0A0A"
FG_MUTED = "#6B6B6B"
BORDER   = "#D0D0D0"
ACCENT   = "#0A0A0A"
ERROR    = "#C0392B"
SUCCESS  = "#1A1A1A"

FONT_TITLE  = ("Georgia", 16, "bold")
FONT_LABEL  = ("Helvetica Neue", 13)
FONT_INPUT  = ("Helvetica Neue", 13)
FONT_BTN    = ("Helvetica Neue", 13, "bold")
FONT_SMALL  = ("Helvetica Neue", 11)
FONT_TABLE  = ("Helvetica Neue", 13)
FONT_HEAD   = ("Helvetica Neue", 12, "bold")

# ── In-memory data store ──────────────────────────────────────────────────────
records = []

COLUMNS = [
    {"key": "name",          "label": "Nombre",                "type": "str",   "required": True},
    {"key": "brazos_r1",     "label": "Brazos Rep 1",          "type": "int"},
    {"key": "brazos_r2",     "label": "Brazos Rep 2",          "type": "int"},
    {"key": "brazos_r3",     "label": "Brazos Rep 3",          "type": "int"},
    {"key": "brazos_avg",    "label": "Brazos Promedio",       "type": "float"},
    {"key": "abdominal_r1",  "label": "Abdominal Rep 1",       "type": "int"},
    {"key": "abdominal_r2",  "label": "Abdominal Rep 2",       "type": "int"},
    {"key": "abdominal_r3",  "label": "Abdominal Rep 3",       "type": "int"},
    {"key": "abdominal_avg", "label": "Abdominal Promedio",    "type": "float"},
    {"key": "salto_c1",      "label": "Salto Intento 1 (cm)",  "type": "float"},
    {"key": "salto_c2",      "label": "Salto Intento 2 (cm)",  "type": "float"},
    {"key": "salto_c3",      "label": "Salto Intento 3 (cm)",  "type": "float"},
    {"key": "salto_max",     "label": "Salto Máximo (cm)",     "type": "float"},
    {"key": "flex_c1",       "label": "Flexibilidad Int 1 (cm)","type": "float"},
    {"key": "flex_c2",       "label": "Flexibilidad Int 2 (cm)","type": "float"},
    {"key": "flex_c3",       "label": "Flexibilidad Int 3 (cm)","type": "float"},
    {"key": "flex_max",      "label": "Flexibilidad Máximo (cm)","type": "float"},
    {"key": "palieres",      "label": "Paliers",               "type": "int"},
    {"key": "vam",           "label": "VAM (km/h)",            "type": "float"},
    {"key": "vo2max",        "label": "VO₂max (ml/kg/min)",    "type": "float"},
]

EXPORT_COLUMNS = [
    {"key": "name",          "label": "ID"},
    {"key": "brazos_avg",    "label": "Brazos Promedio"},
    {"key": "abdominal_avg", "label": "Abdominal Promedio"},
    {"key": "salto_max",     "label": "Salto Máximo (cm)"},
    {"key": "flex_max",      "label": "Flexibilidad Máximo (cm)"},
    {"key": "vam",           "label": "VAM (km/h)"},
    {"key": "vo2max",        "label": "VO₂max (ml/kg/min)"},
]

# ── ID configuration ───────────────────────────────────────────────────────────
ID_TYPES = ["FAE", "PSI", "MEC", "DER", "IND", "SIS", "NEG", "CIV"]

ID_SET = set()
_csv_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "IDs.csv")
if os.path.isfile(_csv_path):
    with open(_csv_path, encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            val = row.get("ID", "").strip().upper()
            if val:
                ID_SET.add(val)

# ── Validators ────────────────────────────────────────────────────────────────
def validate_field(value: str, col: dict):
    label = col["label"]
    t     = col["type"]
    if col["required"] and not value.strip():
        return f"{label} es obligatorio."
    if not value.strip():
        return None
    if t == "int":
        try:
            int(value)
        except ValueError:
            return f"{label} debe ser un número entero (ej: 25)."
    elif t == "float":
        try:
            float(value)
        except ValueError:
            return f"{label} debe ser un número decimal (ej: 1500.50)."
    elif t == "email":
        if not re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", value):
            return f"{label} no tiene un formato válido (ej: a@b.com)."
    elif t == "bool":
        if value.strip().lower() not in ("true", "false", "1", "0", "sí", "si", "no"):
            return f"{label} debe ser: true/false, sí/no."
    return None

def cast_field(value: str, col: dict):
    t = col["type"]
    v = value.strip()
    if not v:
        return None
    if t == "int":   return int(v)
    if t == "float": return float(v)
    if t == "bool":  return v.lower() in ("true", "1", "sí", "si")
    return v

# ── Main window ───────────────────────────────────────────────────────────────
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Gestor de Registros")
        self.configure(bg=BG)
        self.resizable(False, False)
        self.geometry("600x460")

        self._center()
        self._build_home()

    def _center(self):
        self.update_idletasks()
        w, h = 600, 460
        sw = self.winfo_screenwidth()
        sh = self.winfo_screenheight()
        self.geometry(f"{w}x{h}+{(sw-w)//2}+{(sh-h)//2}")

    def _build_home(self):
        for widget in self.winfo_children():
            widget.destroy()

        # ── Header ──
        hdr = tk.Frame(self, bg=BG, pady=0)
        hdr.pack(fill="x", padx=48, pady=(52, 0))

        tk.Label(hdr, text="GESTOR", font=("Georgia", 26, "bold"),
                 bg=BG, fg=FG).pack(side="left")
        tk.Label(hdr, text="  /  registros", font=("Helvetica Neue", 14),
                 bg=BG, fg=FG_MUTED).pack(side="left", pady=(8, 0))

        # ── Divider ──
        sep = tk.Frame(self, bg=BORDER, height=1)
        sep.pack(fill="x", padx=48, pady=(24, 0))

        # ── Subtitle ──
        tk.Label(self,
                 text="Selecciona una acción para continuar.",
                 font=FONT_LABEL, bg=BG, fg=FG_MUTED
                 ).pack(pady=(28, 0))

        # ── Buttons row ──
        row = tk.Frame(self, bg=BG)
        row.pack(pady=40)

        self._big_btn(row, icon="+", label="Crear",
                      cmd=self._open_id_search).pack(side="left", padx=24)
        self._big_btn(row, icon="✎", label="Editar",
                      cmd=self._open_edit).pack(side="left", padx=24)

        # ── Footer ──
        tk.Label(self,
                 text=f"{len(records)} registro(s) almacenado(s)",
                 font=FONT_SMALL, bg=BG, fg=FG_MUTED
                 ).pack(side="bottom", pady=22)

    def _big_btn(self, parent, icon, label, cmd):
        frame = tk.Frame(parent, bg=BG, cursor="hand2")

        btn = tk.Frame(frame, bg=ACCENT, width=140, height=140,
                       relief="flat", bd=0)
        btn.pack_propagate(False)
        btn.pack()

        icon_lbl = tk.Label(btn, text=icon,
                            font=("Georgia", 40), bg=ACCENT, fg=BG)
        icon_lbl.pack(expand=True)

        txt = tk.Label(frame, text=label.upper(),
                       font=("Helvetica Neue", 11, "bold"),
                       bg=BG, fg=FG)
        txt.pack(pady=(12, 0))

        # Hover effects
        def on_enter(e):
            btn.config(bg=FG_MUTED)
            icon_lbl.config(bg=FG_MUTED)
        def on_leave(e):
            btn.config(bg=ACCENT)
            icon_lbl.config(bg=ACCENT)
        def on_click(e): cmd()

        for w in (btn, icon_lbl, frame, txt):
            w.bind("<Enter>", on_enter)
            w.bind("<Leave>", on_leave)
            w.bind("<Button-1>", on_click)

        return frame

    # ── ID SEARCH ────────────────────────────────────────────────────────────
    def _open_id_search(self):
        win = tk.Toplevel(self)
        win.title("Buscar ID")
        win.configure(bg=BG)
        win.resizable(False, False)
        win.geometry("500x380")
        win.transient(self)
        win.grab_set()
        _center_window(win, 500, 380)

        hdr = tk.Frame(win, bg=BG)
        hdr.pack(fill="x", padx=40, pady=(32, 0))
        tk.Label(hdr, text="Buscar ID", font=FONT_TITLE,
                 bg=BG, fg=FG).pack(side="left")

        sep = tk.Frame(win, bg=BORDER, height=1)
        sep.pack(fill="x", padx=40, pady=(16, 24))

        body = tk.Frame(win, bg=BG)
        body.pack(fill="both", expand=True, padx=40)

        # Type label
        tk.Label(body, text="Tipo de ID", font=FONT_LABEL,
                 bg=BG, fg=FG).pack(anchor="w")
        type_var = tk.StringVar(value=ID_TYPES[0])
        type_menu = ttk.Combobox(body, textvariable=type_var,
                                 values=ID_TYPES, state="readonly",
                                 font=FONT_INPUT)
        type_menu.pack(fill="x", ipady=6, pady=(6, 20))

        # Number entry
        tk.Label(body, text="Número (3 dígitos)", font=FONT_LABEL,
                 bg=BG, fg=FG).pack(anchor="w")
        num_var = tk.StringVar()
        num_entry = tk.Entry(body, textvariable=num_var,
                             font=FONT_INPUT, bg=BG2, fg=FG,
                             relief="flat", bd=0,
                             insertbackground=FG,
                             highlightthickness=1,
                             highlightbackground=BORDER,
                             highlightcolor=ACCENT)
        num_entry.pack(fill="x", ipady=10, pady=(6, 24))

        def search():
            raw_type = type_var.get().strip()
            raw_num  = num_var.get().strip()
            if not raw_num:
                return
            full_id = f"USTA{raw_type}{raw_num}".upper()
            if full_id in ID_SET:
                win.destroy()
                self._open_create(prefill_name=full_id)
            else:
                messagebox.showerror(
                    "ID no encontrado",
                    f"Ese ID no existe.",
                    parent=win,
                )

        _solid_btn(body, "Buscar →", search).pack()

    # ── CREATE / EDIT FORM (5 physical tests) ─────────────────────────────
    def _open_create(self, edit_index=None, prefill_name=None):
        is_edit = edit_index is not None
        win = tk.Toplevel(self)
        win.title("Editar Registro" if is_edit else "Nuevo Registro")
        win.configure(bg=BG)
        win.resizable(True, True)
        win.geometry("720x820")
        win.transient(self)
        win.grab_set()
        _center_window(win, 720, 820)

        hdr = tk.Frame(win, bg=BG)
        hdr.pack(fill="x", padx=36, pady=(28, 0))
        tk.Label(hdr, text="Editar Registro" if is_edit else "Nuevo Registro",
                 font=FONT_TITLE, bg=BG, fg=FG).pack(side="left")

        sep = tk.Frame(win, bg=BORDER, height=1)
        sep.pack(fill="x", padx=36, pady=(14, 18))

        canvas = tk.Canvas(win, bg=BG, highlightthickness=0)
        scrollb = ttk.Scrollbar(win, orient="vertical", command=canvas.yview)
        body = tk.Frame(canvas, bg=BG)

        body.bind("<Configure>",
                  lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=body, anchor="nw")
        canvas.configure(yscrollcommand=scrollb.set)

        canvas.pack(side="left", fill="both", expand=True, padx=36)
        scrollb.pack(side="right", fill="y", pady=(0, 64))

        def _on_mousewheel(e):
            canvas.yview_scroll(-1 * (e.delta // 120), "units")
        canvas.bind_all("<MouseWheel>", _on_mousewheel)
        win.protocol("WM_DELETE_WINDOW", lambda: (canvas.unbind_all("<MouseWheel>"), win.destroy()))

        existing = records[edit_index] if is_edit else {}

        def _int(s):
            try: return int(s.strip()) if s.strip() else None
            except: return None
        def _flt(s):
            try: return float(s.strip()) if s.strip() else None
            except: return None

        # ── Name ──────────────────────────────────────────────────────────
        name_var = tk.StringVar()
        if existing.get("name"):
            name_var.set(str(existing["name"]))
        elif prefill_name:
            name_var.set(prefill_name)

        lbl_row = tk.Frame(body, bg=BG)
        lbl_row.pack(fill="x", pady=(0, 4))
        tk.Label(lbl_row, text="Nombre", font=FONT_LABEL,
                 bg=BG, fg=FG).pack(side="left")
        tk.Label(lbl_row, text=" *", font=FONT_SMALL,
                 bg=BG, fg=ERROR).pack(side="left")

        if not is_edit and prefill_name:
            tk.Label(body, text=prefill_name,
                     font=FONT_INPUT, bg=BG2, fg=FG,
                     anchor="w", padx=12,
                     highlightthickness=1,
                     highlightbackground=BORDER
                     ).pack(fill="x", ipady=10, pady=(0, 18))
        else:
            tk.Entry(body, textvariable=name_var,
                     font=FONT_INPUT, bg=BG2, fg=FG,
                     relief="flat", bd=0,
                     insertbackground=FG,
                     highlightthickness=1,
                     highlightbackground=BORDER,
                     highlightcolor=ACCENT
                     ).pack(fill="x", ipady=10, pady=(0, 18))

        # ── Helpers ───────────────────────────────────────────────────────
        def _section_sep(parent):
            tk.Frame(parent, bg=BORDER, height=1).pack(fill="x", pady=(14, 8))

        def _row3(parent, labels, calc_title, is_int):
            """3 small entries side-by-side + a read‑only calc label."""
            row = tk.Frame(parent, bg=BG)
            row.pack(fill="x")
            vars = [tk.StringVar() for _ in range(3)]
            for lbl, var in zip(labels, vars):
                col = tk.Frame(row, bg=BG)
                col.pack(side="left", padx=(0, 10))
                tk.Label(col, text=lbl, font=FONT_SMALL,
                         bg=BG, fg=FG_MUTED).pack(anchor="w")
                tk.Entry(col, textvariable=var,
                         width=5, font=FONT_INPUT,
                         bg=BG2, fg=FG,
                         relief="flat", bd=0,
                         insertbackground=FG,
                         highlightthickness=1,
                         highlightbackground=BORDER,
                         highlightcolor=ACCENT,
                         justify="center"
                         ).pack(ipady=8, pady=(2, 0))
            box = tk.Frame(row, bg=BG)
            box.pack(side="left", padx=(8, 0))
            tk.Label(box, text=calc_title, font=FONT_SMALL,
                     bg=BG, fg=FG_MUTED).pack(anchor="w")
            calc = tk.Label(box, text="—",
                            font=FONT_INPUT, bg=BG2, fg=ACCENT,
                            width=7, anchor="center",
                            highlightthickness=1,
                            highlightbackground=BORDER)
            calc.pack(ipady=8, pady=(2, 0))
            return vars, calc

        # ── Prueba 1: Fuerza de brazos ────────────────────────────────────
        _section_sep(body)
        tk.Label(body, text="Prueba 1: Fuerza de brazos (repeticiones)",
                 font=FONT_HEAD, bg=BG, fg=FG).pack(anchor="w", pady=(0, 6))
        bv, bcalc = _row3(body, ["Rep 1", "Rep 2", "Rep 3"], "Promedio", True)
        if existing.get("brazos_r1") is not None:
            bv[0].set(str(existing["brazos_r1"]))
            bv[1].set(str(existing["brazos_r2"]))
            bv[2].set(str(existing["brazos_r3"]))
        def _upd_brazos(*_):
            vals = [_int(v.get()) for v in bv]
            bcalc.config(text=f"{sum(vals)/3:.1f}" if all(v is not None for v in vals) else "—")
        for v in bv: v.trace_add("write", _upd_brazos)

        # ── Prueba 2: Fuerza abdominal ────────────────────────────────────
        _section_sep(body)
        tk.Label(body, text="Prueba 2: Fuerza abdominal (repeticiones)",
                 font=FONT_HEAD, bg=BG, fg=FG).pack(anchor="w", pady=(0, 6))
        av, acalc = _row3(body, ["Rep 1", "Rep 2", "Rep 3"], "Promedio", True)
        if existing.get("abdominal_r1") is not None:
            av[0].set(str(existing["abdominal_r1"]))
            av[1].set(str(existing["abdominal_r2"]))
            av[2].set(str(existing["abdominal_r3"]))
        def _upd_abdominal(*_):
            vals = [_int(v.get()) for v in av]
            acalc.config(text=f"{sum(vals)/3:.1f}" if all(v is not None for v in vals) else "—")
        for v in av: v.trace_add("write", _upd_abdominal)

        # ── Prueba 3: Salto ───────────────────────────────────────────────
        _section_sep(body)
        tk.Label(body, text="Prueba 3: Fuerza inferior — Salto (cm)",
                 font=FONT_HEAD, bg=BG, fg=FG).pack(anchor="w", pady=(0, 6))
        sv, scalc = _row3(body, ["Intento 1", "Intento 2", "Intento 3"], "Máximo", False)
        if existing.get("salto_c1") is not None:
            sv[0].set(str(existing["salto_c1"]))
            sv[1].set(str(existing["salto_c2"]))
            sv[2].set(str(existing["salto_c3"]))
        def _upd_salto(*_):
            vals = [_flt(v.get()) for v in sv if _flt(v.get()) is not None]
            scalc.config(text=f"{max(vals):.1f}" if vals else "—")
        for v in sv: v.trace_add("write", _upd_salto)

        # ── Prueba 4: Flexibilidad estática ───────────────────────────────
        _section_sep(body)
        tk.Label(body, text="Prueba 4: Flexibilidad estática (cm)",
                 font=FONT_HEAD, bg=BG, fg=FG).pack(anchor="w", pady=(0, 6))
        fv, fcalc = _row3(body, ["Intento 1", "Intento 2", "Intento 3"], "Máximo", False)
        if existing.get("flex_c1") is not None:
            fv[0].set(str(existing["flex_c1"]))
            fv[1].set(str(existing["flex_c2"]))
            fv[2].set(str(existing["flex_c3"]))
        def _upd_flex(*_):
            vals = [_flt(v.get()) for v in fv if _flt(v.get()) is not None]
            fcalc.config(text=f"{max(vals):.1f}" if vals else "—")
        for v in fv: v.trace_add("write", _upd_flex)

        # ── Prueba 5: Resistencia 20m (Course‑Navette) ────────────────────
        _section_sep(body)
        tk.Label(body, text="Prueba 5: Resistencia 20m (Course-Navette)",
                 font=FONT_HEAD, bg=BG, fg=FG).pack(anchor="w", pady=(0, 6))

        row5 = tk.Frame(body, bg=BG)
        row5.pack(fill="x")

        # Palier input
        pc = tk.Frame(row5, bg=BG)
        pc.pack(side="left", padx=(0, 16))
        tk.Label(pc, text="Paliers", font=FONT_SMALL,
                 bg=BG, fg=FG_MUTED).pack(anchor="w")
        pvar = tk.StringVar()
        if existing.get("palieres") is not None:
            pvar.set(str(existing["palieres"]))
        tk.Entry(pc, textvariable=pvar,
                 width=5, font=FONT_INPUT,
                 bg=BG2, fg=FG,
                 relief="flat", bd=0,
                 insertbackground=FG,
                 highlightthickness=1,
                 highlightbackground=BORDER,
                 highlightcolor=ACCENT,
                 justify="center"
                 ).pack(ipady=8, pady=(2, 0))

        # VAM display
        vc = tk.Frame(row5, bg=BG)
        vc.pack(side="left", padx=(0, 16))
        tk.Label(vc, text="VAM (km/h)", font=FONT_SMALL,
                 bg=BG, fg=FG_MUTED).pack(anchor="w")
        vam_lbl = tk.Label(vc, text="—",
                           font=FONT_INPUT, bg=BG2, fg=ACCENT,
                           width=7, anchor="center",
                           highlightthickness=1,
                           highlightbackground=BORDER)
        vam_lbl.pack(ipady=8, pady=(2, 0))

        # VO2max display
        oc = tk.Frame(row5, bg=BG)
        oc.pack(side="left")
        tk.Label(oc, text="VO₂max (ml/kg/min)", font=FONT_SMALL,
                 bg=BG, fg=FG_MUTED).pack(anchor="w")
        vo2_lbl = tk.Label(oc, text="—",
                           font=FONT_INPUT, bg=BG2, fg=ACCENT,
                           width=7, anchor="center",
                           highlightthickness=1,
                           highlightbackground=BORDER)
        vo2_lbl.pack(ipady=8, pady=(2, 0))

        if existing.get("vam") is not None:
            vam_lbl.config(text=f"{existing['vam']:.1f}")
            vo2_lbl.config(text=f"{existing['vo2max']:.1f}")

        def _upd_resistencia(*_):
            try:
                p = int(pvar.get().strip())
                if p > 0:
                    vam = 7.5 + 0.5 * p
                    vo2 = 3.5 * vam
                    vam_lbl.config(text=f"{vam:.1f}")
                    vo2_lbl.config(text=f"{vo2:.1f}")
                    return
            except:
                pass
            vam_lbl.config(text="—")
            vo2_lbl.config(text="—")
        pvar.trace_add("write", _upd_resistencia)

        # ── Spacer to keep content above footer ────────────────────────────
        tk.Frame(body, bg=BG, height=80).pack()

        # ── Footer ────────────────────────────────────────────────────────
        footer = tk.Frame(win, bg=BG, bd=0)
        footer.place(relx=0, rely=1.0, anchor="sw", relwidth=1)

        sep2 = tk.Frame(footer, bg=BORDER, height=1)
        sep2.pack(fill="x")

        btn_row = tk.Frame(footer, bg=BG)
        btn_row.pack(fill="x", padx=36, pady=14)

        def submit():
            name_val = name_var.get().strip()
            if not name_val:
                messagebox.showerror("Error", "Nombre es obligatorio.", parent=win)
                return

            def _store_3(vars, keys, is_int):
                parser = _int if is_int else _flt
                vals = [parser(v.get()) for v in vars]
                d = {}
                any_val = any(v is not None for v in vals)
                d[keys[0]], d[keys[1]], d[keys[2]] = vals if any_val else (None, None, None)
                d[keys[3]] = round(sum(vals)/3, 1) if any_val and all(v is not None for v in vals) else None
                return d

            rec = {"name": name_val}

            rec.update(_store_3(bv, ["brazos_r1","brazos_r2","brazos_r3","brazos_avg"], True))
            rec.update(_store_3(av, ["abdominal_r1","abdominal_r2","abdominal_r3","abdominal_avg"], True))

            for src, keys in [(sv, ["salto_c1","salto_c2","salto_c3","salto_max"]),
                              (fv, ["flex_c1","flex_c2","flex_c3","flex_max"])]:
                vals = [_flt(v.get()) for v in src]
                clean = [x for x in vals if x is not None]
                if clean:
                    rec[keys[0]], rec[keys[1]], rec[keys[2]] = vals
                    rec[keys[3]] = round(max(clean), 1)
                else:
                    for k in keys: rec[k] = None

            pv = _int(pvar.get())
            if pv is not None and pv > 0:
                vam = 7.5 + 0.5 * pv
                rec["palieres"] = pv
                rec["vam"] = round(vam, 1)
                rec["vo2max"] = round(3.5 * vam, 1)
            else:
                rec["palieres"] = rec["vam"] = rec["vo2max"] = None

            if is_edit:
                records[edit_index] = rec
                msg_txt = "Registro actualizado correctamente."
            else:
                if any(r.get("name") == name_val for r in records):
                    messagebox.showerror("Error", "Ese ID ya fue registrado.", parent=win)
                    return
                records.append(rec)
                msg_txt = "Registro creado correctamente."

            win.destroy()
            self._build_home()
            messagebox.showinfo("Éxito", msg_txt, parent=self)

        def cancel():
            win.destroy()

        _outline_btn(btn_row, "Cancelar", cancel, muted=True).pack(side="left")
        _solid_btn(btn_row, "Guardar →", submit).pack(side="right")

# ── EDIT TABLE ───────────────────────────────────────────────────────────
    def _open_edit(self):
        win = tk.Toplevel(self)
        win.title("Editar Registros")
        win.configure(bg=BG)
        win.resizable(True, True)
        win.geometry("680x520")
        win.transient(self)
        win.grab_set()

        _center_window(win, 680, 520)

        # Header
        hdr = tk.Frame(win, bg=BG)
        hdr.pack(fill="x", padx=36, pady=(28, 0))
        tk.Label(hdr, text="Registros", font=FONT_TITLE,
                 bg=BG, fg=FG).pack(side="left")
        count = tk.Label(hdr, text=f"  {len(records)} entrada(s)",
                         font=FONT_SMALL, bg=BG, fg=FG_MUTED)
        count.pack(side="left", pady=(4, 0))

        sep = tk.Frame(win, bg=BORDER, height=1)
        sep.pack(fill="x", padx=36, pady=(14, 10))

        # Scrollable list
        canvas  = tk.Canvas(win, bg=BG, highlightthickness=0)
        scrollb = ttk.Scrollbar(win, orient="vertical", command=canvas.yview)
        body    = tk.Frame(canvas, bg=BG)

        body.bind("<Configure>",
                  lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=body, anchor="nw")
        canvas.configure(yscrollcommand=scrollb.set)

        canvas.pack(side="left", fill="both", expand=True, padx=36, pady=(0, 64))
        scrollb.pack(side="right", fill="y", pady=(0, 64))

        def _on_mousewheel(e):
            canvas.yview_scroll(-1 * (e.delta // 120), "units")
        canvas.bind_all("<MouseWheel>", _on_mousewheel)
        win.protocol("WM_DELETE_WINDOW", lambda: (canvas.unbind_all("<MouseWheel>"), win.destroy()))

        # Column headers
        head = tk.Frame(body, bg=BG2)
        head.pack(fill="x", pady=(0, 6))
        tk.Label(head, text="Nombre", font=FONT_HEAD,
                 bg=BG2, fg=FG, padx=16, pady=10).pack(side="left")
        tk.Label(head, text="Acciones", font=FONT_HEAD,
                 bg=BG2, fg=FG, padx=16, pady=10).pack(side="right")

        def refresh_list():
            # Clear all rows (keep header)
            for w in body.winfo_children():
                if w != head:
                    w.destroy()
            count.config(text=f"  {len(records)} entrada(s)")

            for i, rec in enumerate(records):
                bg_row = BG if i % 2 == 0 else BG2
                row = tk.Frame(body, bg=bg_row,
                               highlightthickness=1,
                               highlightbackground=BORDER)
                row.pack(fill="x", pady=(0, 1))

                name = rec.get("name") or "—"
                tk.Label(row, text=name, font=FONT_TABLE,
                         bg=bg_row, fg=FG, padx=16, pady=14,
                         anchor="w").pack(side="left", fill="x", expand=True)

                btn_box = tk.Frame(row, bg=bg_row)
                btn_box.pack(side="right", padx=10, pady=8)

                def make_edit(idx):
                    def _edit():
                        win.destroy()
                        self._open_create(edit_index=idx)
                    return _edit

                def make_delete(idx):
                    def _delete():
                        if messagebox.askyesno(
                                "Confirmar",
                                f"¿Eliminar '{records[idx].get('name')}'?",
                                parent=win):
                            records.pop(idx)
                            refresh_list()
                            self._build_home()
                    return _delete

                _outline_btn(btn_box, "✎  Editar",
                             make_edit(i), muted=True).pack(side="left", padx=(0, 6))
                _outline_btn(btn_box, "✕  Borrar",
                             make_delete(i), muted=False).pack(side="left")

            if not records:
                tk.Label(body, text="No hay registros todavía.",
                         font=FONT_SMALL, bg=BG, fg=FG_MUTED,
                         pady=24).pack()

        refresh_list()

        # Footer
        sep2 = tk.Frame(win, bg=BORDER, height=1)
        sep2.place(relx=0, rely=1.0, anchor="sw", relwidth=1, y=-52)

        footer = tk.Frame(win, bg=BG)
        footer.place(relx=0, rely=1.0, anchor="sw", relwidth=1)

        tk.Label(footer, text="Haz clic en ✎ Editar para modificar un registro.",
                 font=FONT_SMALL, bg=BG, fg=FG_MUTED,
                 padx=36, pady=14).pack(side="left")

        _outline_btn(footer, "Exportar a Excel →",
                     self._export_to_excel, muted=False).pack(side="right", padx=20, pady=10)

    def _export_to_excel(self):
        if not records:
            messagebox.showwarning("Sin datos", "No hay registros para exportar.", parent=self)
            return

        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            title="Guardar como Excel",
            parent=self,
        )
        if not path:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Registros"

        header_font = Font(name="Helvetica Neue", bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="0A0A0A", end_color="0A0A0A", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin", color="D0D0D0"),
            right=Side(style="thin", color="D0D0D0"),
            top=Side(style="thin", color="D0D0D0"),
            bottom=Side(style="thin", color="D0D0D0"),
        )

        headers = [col["label"] for col in EXPORT_COLUMNS]
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        for r, rec in enumerate(records, 2):
            for c, col in enumerate(EXPORT_COLUMNS, 1):
                val = rec.get(col["key"])
                if val is None:
                    val = ""
                cell = ws.cell(row=r, column=c, value=val)
                cell.font = Font(name="Helvetica Neue", size=10)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border

        for c in range(1, len(EXPORT_COLUMNS) + 1):
            ws.column_dimensions[chr(64 + c)].width = 22

        wb.save(path)
        messagebox.showinfo(
            "Exportado",
            f"Se exportaron {len(records)} registro(s) a:\n{path}",
            parent=self,
        )

# ── Helpers ───────────────────────────────────────────────────────────────────
def _center_window(win, w, h):
    win.update_idletasks()
    sw = win.winfo_screenwidth()
    sh = win.winfo_screenheight()
    win.geometry(f"{w}x{h}+{(sw-w)//2}+{(sh-h)//2}")

def _solid_btn(parent, text, cmd):
    btn = tk.Label(parent, text=text, font=FONT_BTN,
                   bg=ACCENT, fg=BG, padx=28, pady=12,
                   cursor="hand2")
    btn.bind("<Button-1>", lambda e: cmd())
    btn.bind("<Enter>", lambda e: btn.config(bg=FG_MUTED))
    btn.bind("<Leave>", lambda e: btn.config(bg=ACCENT))
    return btn

def _outline_btn(parent, text, cmd, muted=False):
    clr = FG_MUTED if muted else FG
    btn = tk.Label(parent, text=text, font=FONT_BTN,
                   bg=BG, fg=clr, padx=22, pady=11,
                   highlightthickness=1,
                   highlightbackground=BORDER,
                   cursor="hand2")
    btn.bind("<Button-1>", lambda e: cmd())
    btn.bind("<Enter>", lambda e: btn.config(bg=BG2))
    btn.bind("<Leave>", lambda e: btn.config(bg=BG))
    return btn

# ── Entry point ───────────────────────────────────────────────────────────────
if __name__ == "__main__":
    app = App()
    app.mainloop()
