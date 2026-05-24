import sqlite3, csv, os, io, json, sys
from flask import Flask, render_template, request, jsonify, send_file, redirect
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

app = Flask(__name__)
EDIT_PASSWORD = "USTAHenriDIDON"
BASE_DIR = os.path.dirname(__file__)
ID_TYPES = ["FAE", "PSI", "MEC", "DER", "IND", "SIS", "NEG", "CIV"]

# ── Load valid IDs ──
ID_SET = set()
_csv_path = os.path.join(BASE_DIR, "IDs.csv")
if os.path.isfile(_csv_path):
    with open(_csv_path, encoding="utf-8") as f:
        for row in csv.DictReader(f):
            val = row.get("ID", "").strip().upper()
            if val:
                ID_SET.add(val)

# ── Database backend ──
PG_URL = (os.environ.get("POSTGRES_URL")
          or os.environ.get("DATABASE_URL")
          or os.environ.get("PRISMA_DATABASE_URL")
          or os.environ.get("POSTGRES_PRISMA_URL"))

def _pg_url():
    url = PG_URL
    if url and "sslmode" not in url:
        url += ("&" if "?" in url else "?") + "sslmode=require"
    return url

def _pg_conn():
    import psycopg2
    import psycopg2.extras
    return psycopg2.connect(_pg_url())

if PG_URL:
    sys.stderr.write(f"[DB] PostgreSQL detected via env var\n")
    import psycopg2
    import psycopg2.extras

    def db_init():
        with _pg_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS registros (
                        id SERIAL PRIMARY KEY,
                        name TEXT NOT NULL,
                        brazos_r1 INTEGER, brazos_r2 INTEGER, brazos_r3 INTEGER, brazos_avg DOUBLE PRECISION,
                        abdominal_r1 INTEGER, abdominal_r2 INTEGER, abdominal_r3 INTEGER, abdominal_avg DOUBLE PRECISION,
                        salto_c1 DOUBLE PRECISION, salto_c2 DOUBLE PRECISION, salto_c3 DOUBLE PRECISION, salto_max DOUBLE PRECISION,
                        flex_c1 DOUBLE PRECISION, flex_c2 DOUBLE PRECISION, flex_c3 DOUBLE PRECISION, flex_max DOUBLE PRECISION,
                        palieres INTEGER, vam DOUBLE PRECISION, vo2max DOUBLE PRECISION
                    )
                """)
            conn.commit()

    def db_all():
        with _pg_conn() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute("SELECT * FROM registros ORDER BY id")
                return [dict(r) for r in cur.fetchall()]

    def db_get(rid):
        with _pg_conn() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute("SELECT * FROM registros WHERE id = %s", (rid,))
                row = cur.fetchone()
                return dict(row) if row else None

    def db_create(data):
        with _pg_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    INSERT INTO registros (
                        name, brazos_r1, brazos_r2, brazos_r3, brazos_avg,
                        abdominal_r1, abdominal_r2, abdominal_r3, abdominal_avg,
                        salto_c1, salto_c2, salto_c3, salto_max,
                        flex_c1, flex_c2, flex_c3, flex_max,
                        palieres, vam, vo2max
                    ) VALUES (%s,%s,%s,%s,%s, %s,%s,%s,%s,%s, %s,%s,%s,%s,%s, %s,%s,%s,%s,%s)
                    RETURNING id
                """, (
                    data["name"],
                    data.get("brazos_r1"), data.get("brazos_r2"), data.get("brazos_r3"), data.get("brazos_avg"),
                    data.get("abdominal_r1"), data.get("abdominal_r2"), data.get("abdominal_r3"), data.get("abdominal_avg"),
                    data.get("salto_c1"), data.get("salto_c2"), data.get("salto_c3"), data.get("salto_max"),
                    data.get("flex_c1"), data.get("flex_c2"), data.get("flex_c3"), data.get("flex_max"),
                    data.get("palieres"), data.get("vam"), data.get("vo2max"),
                ))
                row_id = cur.fetchone()[0]
            conn.commit()
            return row_id

    def db_update(rid, data):
        with _pg_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    UPDATE registros SET
                        name=%s, brazos_r1=%s, brazos_r2=%s, brazos_r3=%s, brazos_avg=%s,
                        abdominal_r1=%s, abdominal_r2=%s, abdominal_r3=%s, abdominal_avg=%s,
                        salto_c1=%s, salto_c2=%s, salto_c3=%s, salto_max=%s,
                        flex_c1=%s, flex_c2=%s, flex_c3=%s, flex_max=%s,
                        palieres=%s, vam=%s, vo2max=%s
                    WHERE id=%s
                """, (
                    data["name"],
                    data.get("brazos_r1"), data.get("brazos_r2"), data.get("brazos_r3"), data.get("brazos_avg"),
                    data.get("abdominal_r1"), data.get("abdominal_r2"), data.get("abdominal_r3"), data.get("abdominal_avg"),
                    data.get("salto_c1"), data.get("salto_c2"), data.get("salto_c3"), data.get("salto_max"),
                    data.get("flex_c1"), data.get("flex_c2"), data.get("flex_c3"), data.get("flex_max"),
                    data.get("palieres"), data.get("vam"), data.get("vo2max"),
                    rid
                ))
            conn.commit()

    def db_delete(rid):
        with _pg_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("DELETE FROM registros WHERE id = %s", (rid,))
            conn.commit()

else:
    DB_PATH = os.path.join(BASE_DIR, "database.db")

    def db_init():
        conn = sqlite3.connect(DB_PATH)
        conn.executescript("""
            CREATE TABLE IF NOT EXISTS registros (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                brazos_r1 INTEGER, brazos_r2 INTEGER, brazos_r3 INTEGER, brazos_avg REAL,
                abdominal_r1 INTEGER, abdominal_r2 INTEGER, abdominal_r3 INTEGER, abdominal_avg REAL,
                salto_c1 REAL, salto_c2 REAL, salto_c3 REAL, salto_max REAL,
                flex_c1 REAL, flex_c2 REAL, flex_c3 REAL, flex_max REAL,
                palieres INTEGER, vam REAL, vo2max REAL
            )
        """)
        conn.commit()
        conn.close()

    def db_all():
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        rows = conn.execute("SELECT * FROM registros ORDER BY id").fetchall()
        conn.close()
        return [dict(r) for r in rows]

    def db_get(rid):
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        row = conn.execute("SELECT * FROM registros WHERE id = ?", (rid,)).fetchone()
        conn.close()
        return dict(row) if row else None

    def db_create(data):
        conn = sqlite3.connect(DB_PATH)
        dup = conn.execute("SELECT id FROM registros WHERE name = ?", (data["name"],)).fetchone()
        if dup:
            conn.close()
            return None
        conn.execute("""INSERT INTO registros (
            name, brazos_r1, brazos_r2, brazos_r3, brazos_avg,
            abdominal_r1, abdominal_r2, abdominal_r3, abdominal_avg,
            salto_c1, salto_c2, salto_c3, salto_max,
            flex_c1, flex_c2, flex_c3, flex_max,
            palieres, vam, vo2max
        ) VALUES (?,?,?,?,?, ?,?,?,?,?, ?,?,?,?,?, ?,?,?,?,?)""", (
            data["name"],
            data.get("brazos_r1"), data.get("brazos_r2"), data.get("brazos_r3"), data.get("brazos_avg"),
            data.get("abdominal_r1"), data.get("abdominal_r2"), data.get("abdominal_r3"), data.get("abdominal_avg"),
            data.get("salto_c1"), data.get("salto_c2"), data.get("salto_c3"), data.get("salto_max"),
            data.get("flex_c1"), data.get("flex_c2"), data.get("flex_c3"), data.get("flex_max"),
            data.get("palieres"), data.get("vam"), data.get("vo2max"),
        ))
        conn.commit()
        row_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
        conn.close()
        return row_id

    def db_update(rid, data):
        conn = sqlite3.connect(DB_PATH)
        conn.execute("""UPDATE registros SET
            name=?, brazos_r1=?, brazos_r2=?, brazos_r3=?, brazos_avg=?,
            abdominal_r1=?, abdominal_r2=?, abdominal_r3=?, abdominal_avg=?,
            salto_c1=?, salto_c2=?, salto_c3=?, salto_max=?,
            flex_c1=?, flex_c2=?, flex_c3=?, flex_max=?,
            palieres=?, vam=?, vo2max=?
            WHERE id=?""", (
            data["name"],
            data.get("brazos_r1"), data.get("brazos_r2"), data.get("brazos_r3"), data.get("brazos_avg"),
            data.get("abdominal_r1"), data.get("abdominal_r2"), data.get("abdominal_r3"), data.get("abdominal_avg"),
            data.get("salto_c1"), data.get("salto_c2"), data.get("salto_c3"), data.get("salto_max"),
            data.get("flex_c1"), data.get("flex_c2"), data.get("flex_c3"), data.get("flex_max"),
            data.get("palieres"), data.get("vam"), data.get("vo2max"),
            rid
        ))
        conn.commit()
        conn.close()

    def db_delete(rid):
        conn = sqlite3.connect(DB_PATH)
        conn.execute("DELETE FROM registros WHERE id = ?", (rid,))
        conn.commit()
        conn.close()

db_init()

# ── Auth helpers ──

def check_edit_auth():
    return request.headers.get("X-Edit-Password", "") == EDIT_PASSWORD

def edit_required(fn):
    from functools import wraps
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if not check_edit_auth():
            return jsonify({"error": "No autorizado"}), 403
        return fn(*args, **kwargs)
    return wrapper

# ── Pages ──

@app.route("/")
def index():
    return render_template("index.html")

@app.route("/buscar")
def buscar():
    return render_template("buscar.html", types=ID_TYPES)

@app.route("/crear")
def crear():
    return render_template("crear.html", name=request.args.get("name", ""))

@app.route("/editar")
def editar():
    return render_template("editar.html")

@app.route("/editar/<int:rid>")
def editar_form(rid):
    return render_template("crear.html", edit_id=rid)

# ── Auth API ──

@app.route("/api/auth", methods=["POST"])
def api_auth():
    data = request.json or {}
    return jsonify({"ok": data.get("password", "") == EDIT_PASSWORD})

# ── Record API ──

@app.route("/api/buscar")
def api_buscar():
    tipo = request.args.get("tipo", "").strip()
    numero = request.args.get("numero", "").strip()
    if not numero:
        return jsonify({"found": False})
    full_id = f"USTA{tipo}{numero}".upper()
    return jsonify({"found": full_id in ID_SET, "id": full_id})

@app.route("/api/records", methods=["GET"])
def api_get_records():
    return jsonify(db_all())

@app.route("/api/records/<int:rid>", methods=["GET"])
def api_get_record(rid):
    row = db_get(rid)
    if row is None:
        return jsonify({"error": "No encontrado"}), 404
    return jsonify(row)

@app.route("/api/records", methods=["POST"])
def api_create_record():
    data = request.json
    row_id = db_create(data)
    if row_id is None:
        return jsonify({"error": "Ese ID ya fue registrado."}), 409
    return jsonify({"ok": True, "id": row_id}), 201

@app.route("/api/records/<int:rid>", methods=["PUT"])
@edit_required
def api_update_record(rid):
    data = request.json
    db_update(rid, data)
    return jsonify({"ok": True})

@app.route("/api/records/<int:rid>", methods=["DELETE"])
@edit_required
def api_delete_record(rid):
    db_delete(rid)
    return jsonify({"ok": True})

# ── Export ──

EXPORT_FIELDS = [
    ("name", "ID"),
    ("brazos_avg", "Brazos Promedio"),
    ("abdominal_avg", "Abdominal Promedio"),
    ("salto_max", "Salto M\u00e1ximo (cm)"),
    ("flex_max", "Flexibilidad M\u00e1ximo (cm)"),
    ("vam", "VAM (km/h)"),
    ("vo2max", "VO\u2082max (ml/kg/min)"),
]

@app.route("/api/export")
@edit_required
def api_export():
    rows = db_all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Registros"

    hfont = Font(name="Helvetica Neue", bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="0A0A0A", end_color="0A0A0A", fill_type="solid")
    halign = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0"),
    )

    for c, (_, label) in enumerate(EXPORT_FIELDS, 1):
        cell = ws.cell(row=1, column=c, value=label)
        cell.font = hfont
        cell.fill = hfill
        cell.alignment = halign
        cell.border = border

    for r, row in enumerate(rows, 2):
        for c, (key, _) in enumerate(EXPORT_FIELDS, 1):
            val = row[key] if row[key] is not None else ""
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = Font(name="Helvetica Neue", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

    for c in range(1, len(EXPORT_FIELDS) + 1):
        ws.column_dimensions[chr(64 + c)].width = 24

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name="registros.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

def _get_local_ip():
    import socket
    s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
    try:
        s.connect(("8.8.8.8", 80))
        ip = s.getsockname()[0]
    except Exception:
        ip = "127.0.0.1"
    finally:
        s.close()
    return ip

if __name__ == "__main__":
    print(f"Edit password: {EDIT_PASSWORD}")
    print(f"DB backend: {'PostgreSQL' if PG_URL else 'SQLite'}")
    print(f"Servidor local: http://localhost:5000")
    print(f"\nRed local: http://{_get_local_ip()}:5000")
    app.run(host="0.0.0.0", port=5000, debug=True)
