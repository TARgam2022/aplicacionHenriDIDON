import sqlite3, csv, os, io
from flask import Flask, render_template, request, jsonify, send_file, redirect
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

app = Flask(__name__)

DB_PATH = os.path.join(os.path.dirname(__file__), "database.db")

ID_TYPES = ["FAE", "PSI", "MEC", "DER", "IND", "SIS", "NEG", "CIV"]

ID_SET = set()
_csv_path = os.path.join(os.path.dirname(__file__), "IDs.csv")
if os.path.isfile(_csv_path):
    with open(_csv_path, encoding="utf-8") as f:
        for row in csv.DictReader(f):
            val = row.get("ID", "").strip().upper()
            if val:
                ID_SET.add(val)

def is_local():
    if request.headers.get("X-Forwarded-For"):
        return False
    return request.remote_addr in ("127.0.0.1", "::1", "localhost")

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_db()
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS registros (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            brazos_r1 INTEGER, brazos_r2 INTEGER, brazos_r3 INTEGER, brazos_avg REAL,
            abdominal_r1 INTEGER, abdominal_r2 INTEGER, abdominal_r3 INTEGER, abdominal_avg REAL,
            salto_c1 REAL, salto_c2 REAL, salto_c3 REAL, salto_max REAL,
            flex_c1 REAL, flex_c2 REAL, flex_c3 REAL, flex_max REAL,
            palieres INTEGER, vam REAL, vo2max REAL
        )
    """)
    conn.commit()
    conn.close()

init_db()

# ── Pages ──────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return render_template("index.html", admin=is_local())

@app.route("/buscar")
def buscar():
    return render_template("buscar.html", types=ID_TYPES)

@app.route("/crear")
def crear():
    return render_template("crear.html", name=request.args.get("name", ""), admin=is_local())

@app.route("/editar")
def editar():
    if not is_local():
        return redirect("/")
    return render_template("editar.html")

@app.route("/editar/<int:rid>")
def editar_form(rid):
    if not is_local():
        return redirect("/")
    return render_template("crear.html", edit_id=rid, admin=True)

# ── API ────────────────────────────────────────────────────────────────────

@app.route("/api/buscar")
def api_buscar():
    tipo = request.args.get("tipo", "").strip()
    numero = request.args.get("numero", "").strip()
    if not numero:
        return jsonify({"found": False})
    full_id = f"USTA{tipo}{numero}".upper()
    return jsonify({"found": full_id in ID_SET, "id": full_id})

@app.route("/api/records", methods=["GET"])
def api_get_records():
    conn = get_db()
    rows = conn.execute("SELECT * FROM registros ORDER BY id").fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route("/api/records/<int:rid>", methods=["GET"])
def api_get_record(rid):
    conn = get_db()
    row = conn.execute("SELECT * FROM registros WHERE id = ?", (rid,)).fetchone()
    conn.close()
    if row is None:
        return jsonify({"error": "No encontrado"}), 404
    return jsonify(dict(row))

@app.route("/api/records", methods=["POST"])
def api_create_record():
    data = request.json
    conn = get_db()
    dup = conn.execute("SELECT id FROM registros WHERE name = ?", (data["name"],)).fetchone()
    if dup:
        conn.close()
        return jsonify({"error": "Ese ID ya fue registrado."}), 409
    conn.execute("""INSERT INTO registros (
        name, brazos_r1, brazos_r2, brazos_r3, brazos_avg,
        abdominal_r1, abdominal_r2, abdominal_r3, abdominal_avg,
        salto_c1, salto_c2, salto_c3, salto_max,
        flex_c1, flex_c2, flex_c3, flex_max,
        palieres, vam, vo2max
    ) VALUES (?,?,?,?,?, ?,?,?,?, ?,?,?,?, ?,?,?,?, ?,?,?)""", (
        data["name"],
        data.get("brazos_r1"), data.get("brazos_r2"), data.get("brazos_r3"), data.get("brazos_avg"),
        data.get("abdominal_r1"), data.get("abdominal_r2"), data.get("abdominal_r3"), data.get("abdominal_avg"),
        data.get("salto_c1"), data.get("salto_c2"), data.get("salto_c3"), data.get("salto_max"),
        data.get("flex_c1"), data.get("flex_c2"), data.get("flex_c3"), data.get("flex_max"),
        data.get("palieres"), data.get("vam"), data.get("vo2max"),
    ))
    conn.commit()
    row_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
    conn.close()
    return jsonify({"ok": True, "id": row_id}), 201

@app.route("/api/records/<int:rid>", methods=["PUT"])
def api_update_record(rid):
    if not is_local():
        return jsonify({"error": "No autorizado"}), 403
    data = request.json
    conn = get_db()
    conn.execute("""UPDATE registros SET
        name=?, brazos_r1=?, brazos_r2=?, brazos_r3=?, brazos_avg=?,
        abdominal_r1=?, abdominal_r2=?, abdominal_r3=?, abdominal_avg=?,
        salto_c1=?, salto_c2=?, salto_c3=?, salto_max=?,
        flex_c1=?, flex_c2=?, flex_c3=?, flex_max=?,
        palieres=?, vam=?, vo2max=?
        WHERE id=?""", (
        data["name"],
        data.get("brazos_r1"), data.get("brazos_r2"), data.get("brazos_r3"), data.get("brazos_avg"),
        data.get("abdominal_r1"), data.get("abdominal_r2"), data.get("abdominal_r3"), data.get("abdominal_avg"),
        data.get("salto_c1"), data.get("salto_c2"), data.get("salto_c3"), data.get("salto_max"),
        data.get("flex_c1"), data.get("flex_c2"), data.get("flex_c3"), data.get("flex_max"),
        data.get("palieres"), data.get("vam"), data.get("vo2max"),
        rid
    ))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})

@app.route("/api/records/<int:rid>", methods=["DELETE"])
def api_delete_record(rid):
    if not is_local():
        return jsonify({"error": "No autorizado"}), 403
    conn = get_db()
    conn.execute("DELETE FROM registros WHERE id = ?", (rid,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})

# ── Export ─────────────────────────────────────────────────────────────────

EXPORT_FIELDS = [
    ("name", "ID"),
    ("brazos_avg", "Brazos Promedio"),
    ("abdominal_avg", "Abdominal Promedio"),
    ("salto_max", "Salto Máximo (cm)"),
    ("flex_max", "Flexibilidad Máximo (cm)"),
    ("vam", "VAM (km/h)"),
    ("vo2max", "VO\u2082max (ml/kg/min)"),
]

@app.route("/api/export")
def api_export():
    if not is_local():
        return jsonify({"error": "No autorizado"}), 403
    conn = get_db()
    rows = conn.execute("SELECT * FROM registros ORDER BY id").fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Registros"

    hfont = Font(name="Helvetica Neue", bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="0A0A0A", end_color="0A0A0A", fill_type="solid")
    halign = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0"),
    )

    for c, (_, label) in enumerate(EXPORT_FIELDS, 1):
        cell = ws.cell(row=1, column=c, value=label)
        cell.font = hfont
        cell.fill = hfill
        cell.alignment = halign
        cell.border = border

    for r, row in enumerate(rows, 2):
        for c, (key, _) in enumerate(EXPORT_FIELDS, 1):
            val = row[key] if row[key] is not None else ""
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = Font(name="Helvetica Neue", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

    for c in range(1, len(EXPORT_FIELDS) + 1):
        ws.column_dimensions[chr(64 + c)].width = 24

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name="registros.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

def _get_local_ip():
    import socket
    s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
    try:
        s.connect(("8.8.8.8", 80))
        ip = s.getsockname()[0]
    except Exception:
        ip = "127.0.0.1"
    finally:
        s.close()
    return ip

if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--public", action="store_true",
                        help="Exponer el servidor p\u00fablicamente con ngrok")
    parser.add_argument("--token", type=str, default="",
                        help="Token de autenticaci\u00f3n de ngrok")
    args = parser.parse_args()

    print("Servidor local: http://localhost:5000")

    if args.public:
        try:
            from pyngrok import ngrok, conf

            if args.token:
                conf.get_default().auth_token = args.token

            tunnel = ngrok.connect(5000, bind_tls=True)
            print(f"P\u00fablico: {tunnel.public_url}")
            print("Cualquier persona con ese enlace puede crear registros.")
        except Exception as e:
            print(f"Error al iniciar ngrok: {e}")
            print("Aseg\u00farate de tener ngrok configurado:")
            print("  1. Reg\u00edstrate en https://ngrok.com")
            print("  2. Usa: python app-server.py --public --token TU_TOKEN")

    print(f"\nRed local: http://{_get_local_ip()}:5000")
    app.run(host="0.0.0.0", port=5000, debug=True)
